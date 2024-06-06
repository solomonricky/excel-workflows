import openpyxl
import sys
import os

def compare_excel_files(new_file_path, old_file_path):
    wb_new = openpyxl.load_workbook(new_file_path)
    wb_old = openpyxl.load_workbook(old_file_path)
    
    sheet_new = wb_new.active
    sheet_old = wb_old.active

    changes = []
    
    for row in sheet_new.iter_rows():
        for cell in row:
            old_value = sheet_old.cell(row=cell.row, column=cell.column).value
            new_value = cell.value
            if old_value != new_value:
                changes.append({
                    "row": cell.row,
                    "column": cell.column,
                    "old_value": old_value,
                    "new_value": new_value
                })

    with open('changes.txt', 'w') as f:
        for change in changes:
            f.write(f"Row {change['row']}, Column {change['column']} changed from '{change['old_value']}' to '{change['new_value']}'\n")
    
    if changes:
        with open(os.environ['GITHUB_ENV'], 'a') as env_file:
            env_file.write("changes_detected=true\n")
    else:
        with open(os.environ['GITHUB_ENV'], 'a') as env_file:
            env_file.write("changes_detected=false\n")

if __name__ == "__main__":
    new_file_path = sys.argv[1]
    old_file_path = sys.argv[2]
    
    compare_excel_files(new_file_path, old_file_path)
